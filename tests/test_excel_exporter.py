"""Tests for src/excel_exporter.py — amount formatting and Excel generation."""

import io

import pytest
from openpyxl import load_workbook

from excel_exporter import _fmt_amount, build_monthly_excel


# ---------------------------------------------------------------------------
# _fmt_amount
# ---------------------------------------------------------------------------

class TestFmtAmount:
    def test_none_returns_empty(self):
        assert _fmt_amount(None) == ""

    def test_zero(self):
        assert _fmt_amount(0) == "0,00"

    def test_positive(self):
        # French format: space as thousands separator, comma as decimal
        result = _fmt_amount(1234.56)
        assert result == "1 234,56"

    def test_negative(self):
        result = _fmt_amount(-50.0)
        assert result == "-50,00"

    def test_large_number(self):
        result = _fmt_amount(1000000.0)
        assert result == "1 000 000,00"

    def test_small_decimal(self):
        result = _fmt_amount(0.99)
        assert result == "0,99"


# ---------------------------------------------------------------------------
# build_monthly_excel
# ---------------------------------------------------------------------------

def _sample_invoices():
    return [
        {
            "invoice_date": "2025-03-01",
            "received_at": "2025-03-02T10:00:00Z",
            "sender": "billing@acme.com",
            "supplier": "Acme Corp",
            "filename": "2025-03-01_acme_facture.pdf",
            "drive_web_link": "https://onedrive.example/file1",
            "currency": "EUR",
            "amount_ht": 100.0,
            "amount_tva": 20.0,
            "amount_ttc": 120.0,
        },
        {
            "invoice_date": "2025-03-15",
            "received_at": "2025-03-16T08:00:00Z",
            "sender": "noreply@bigcorp.fr",
            "supplier": "BigCorp",
            "filename": "2025-03-15_bigcorp_invoice.pdf",
            "drive_web_link": "https://onedrive.example/file2",
            "currency": "EUR",
            "amount_ht": 200.0,
            "amount_tva": 40.0,
            "amount_ttc": 240.0,
        },
    ]


class TestBuildMonthlyExcel:
    def test_returns_bytes(self):
        result = build_monthly_excel(_sample_invoices(), 2025, 3)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx(self):
        result = build_monthly_excel(_sample_invoices(), 2025, 3)
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) >= 1

    def test_sheet_name_matches_month(self):
        result = build_monthly_excel(_sample_invoices(), 2025, 3)
        wb = load_workbook(io.BytesIO(result))
        assert wb.sheetnames[0] == "Mars 2025"

    def test_header_row(self):
        result = build_monthly_excel(_sample_invoices(), 2025, 3)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 11)]
        assert "Date" in headers
        assert "Fournisseur" in headers
        assert "HT" in headers
        assert "TTC" in headers

    def test_data_row_count(self):
        invoices = _sample_invoices()
        result = build_monthly_excel(invoices, 2025, 3)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # Row 1 = header, rows 2..N+1 = data
        data_rows = 0
        for row in range(2, 2 + len(invoices)):
            if ws.cell(row=row, column=1).value is not None:
                data_rows += 1
        assert data_rows == len(invoices)

    def test_empty_invoices(self):
        result = build_monthly_excel([], 2025, 1)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        # Should still have header row
        assert ws.cell(row=1, column=1).value == "Date"

    def test_invoice_without_amounts(self):
        invoices = [{
            "received_at": "2025-04-01T00:00:00Z",
            "sender": "a@b.com",
            "filename": "file.pdf",
            "drive_web_link": "",
            "amount_ht": None,
            "amount_tva": None,
            "amount_ttc": None,
            "currency": None,
        }]
        result = build_monthly_excel(invoices, 2025, 4)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_frozen_panes(self):
        result = build_monthly_excel(_sample_invoices(), 2025, 3)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.freeze_panes == "A2"

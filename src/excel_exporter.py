"""
Monthly invoice summary Excel exporter.

Builds a styled .xlsx workbook in memory from a list of invoice records
and returns the raw bytes, ready to be uploaded to OneDrive.

Sheet layout:
  - Header row (frozen)
  - One data row per invoice with amounts
  - Blank separator row
  - Per-supplier summary block
  - Grand total row
"""

import io
import logging
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from onedrive_uploader import _sender_to_label
from utils import MONTH_NAMES_FR

logger = logging.getLogger(__name__)

# Colours
HEADER_BG    = "2C5F8A"
HEADER_FG    = "FFFFFF"
SUMMARY_BG   = "E8F0F7"  # light blue for summary section header
TOTAL_BG     = "2C5F8A"  # same as main header for grand total row
TOTAL_FG     = "FFFFFF"

# (header label, min width, max width)
COLUMNS = [
    ("Date",              12, 14),
    ("Fournisseur",       20, 35),
    ("Email expéditeur",  28, 42),
    ("Nom du fichier",    35, 55),
    ("HT",                12, 16),
    ("TVA",               12, 16),
    ("TTC",               12, 16),
    ("Devise",             8,  8),
    ("Lien Drive",        18, 18),
    ("Période",           12, 12),
]

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(bottom=_THIN)


def _fmt_amount(value: float | None) -> str:
    """Format a float amount as a string, or empty string if None."""
    if value is None:
        return ""
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def build_monthly_excel(invoices: list[dict], year: int, month: int) -> bytes:
    """
    Build a monthly invoice summary workbook.

    Args:
        invoices: List of invoice dicts from db.get_unreported_invoices().
        year:     Report year.
        month:    Report month (1-12).

    Returns:
        Raw .xlsx bytes.
    """
    month_label = f"{MONTH_NAMES_FR[month]} {year}"
    period_str = f"{month:02d}/{year}"

    wb = Workbook()
    ws = wb.active
    ws.title = month_label

    # -----------------------------------------------------------------------
    # Styles
    # -----------------------------------------------------------------------
    header_font     = Font(bold=True, color=HEADER_FG, name="Calibri", size=11)
    header_fill     = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    header_align    = Alignment(horizontal="center", vertical="center")

    row_font        = Font(name="Calibri", size=10)
    link_font       = Font(name="Calibri", size=10, color="0563C1", underline="single")
    num_font        = Font(name="Calibri", size=10)
    center          = Alignment(horizontal="center", vertical="center")
    left            = Alignment(horizontal="left",   vertical="center")
    right           = Alignment(horizontal="right",  vertical="center")

    sum_hdr_font    = Font(bold=True, name="Calibri", size=10, color="2C5F8A")
    sum_hdr_fill    = PatternFill(fill_type="solid", fgColor=SUMMARY_BG)
    sum_row_font    = Font(name="Calibri", size=10)
    sum_bold_font   = Font(bold=True, name="Calibri", size=10)

    total_font      = Font(bold=True, color=TOTAL_FG, name="Calibri", size=10)
    total_fill      = PatternFill(fill_type="solid", fgColor=TOTAL_BG)

    # -----------------------------------------------------------------------
    # Header row
    # -----------------------------------------------------------------------
    for col_idx, (label, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20

    # -----------------------------------------------------------------------
    # Data rows
    # -----------------------------------------------------------------------
    supplier_totals: dict[str, dict] = defaultdict(lambda: {"count": 0, "ht": 0.0, "tva": 0.0, "ttc": 0.0, "has_amounts": False})

    for row_idx, inv in enumerate(invoices, start=2):
        raw_date = inv.get("invoice_date") or inv.get("received_at", "")
        try:
            dt = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
            date_str = dt.strftime("%d/%m/%Y")
        except Exception:
            date_str = raw_date

        sender: str = inv.get("sender", "")
        supplier: str = inv.get("supplier") or _sender_to_label(sender).capitalize()
        filename: str = inv.get("filename", "")
        drive_link: str = inv.get("drive_web_link", "")
        currency: str = inv.get("currency") or "EUR"

        amount_ht  = inv.get("amount_ht")
        amount_tva = inv.get("amount_tva")
        amount_ttc = inv.get("amount_ttc")

        st = supplier_totals[supplier]
        st["count"] += 1
        if amount_ht is not None:
            st["ht"] += amount_ht
            st["has_amounts"] = True
        if amount_tva is not None:
            st["tva"] += amount_tva
            st["has_amounts"] = True
        if amount_ttc is not None:
            st["ttc"] += amount_ttc
            st["has_amounts"] = True

        c = ws.cell(row=row_idx, column=1, value=date_str)
        c.font = row_font; c.alignment = center; c.border = _BORDER

        c = ws.cell(row=row_idx, column=2, value=supplier)
        c.font = row_font; c.alignment = left; c.border = _BORDER

        c = ws.cell(row=row_idx, column=3, value=sender)
        c.font = row_font; c.alignment = left; c.border = _BORDER

        c = ws.cell(row=row_idx, column=4, value=filename)
        c.font = row_font; c.alignment = left; c.border = _BORDER

        c = ws.cell(row=row_idx, column=5, value=_fmt_amount(amount_ht))
        c.font = num_font; c.alignment = right; c.border = _BORDER

        c = ws.cell(row=row_idx, column=6, value=_fmt_amount(amount_tva))
        c.font = num_font; c.alignment = right; c.border = _BORDER

        c = ws.cell(row=row_idx, column=7, value=_fmt_amount(amount_ttc))
        c.font = num_font; c.alignment = right; c.border = _BORDER

        c = ws.cell(row=row_idx, column=8, value=currency)
        c.font = row_font; c.alignment = center; c.border = _BORDER

        if drive_link:
            c = ws.cell(row=row_idx, column=9, value=f'=HYPERLINK("{drive_link}","Ouvrir")')
            c.font = link_font
        else:
            c = ws.cell(row=row_idx, column=9, value="—")
            c.font = row_font
        c.alignment = center; c.border = _BORDER

        c = ws.cell(row=row_idx, column=10, value=period_str)
        c.font = row_font; c.alignment = center; c.border = _BORDER

        ws.row_dimensions[row_idx].height = 16

    # -----------------------------------------------------------------------
    # Per-supplier summary block
    # -----------------------------------------------------------------------
    blank_row = len(invoices) + 2
    ws.row_dimensions[blank_row].height = 10

    summary_header_row = blank_row + 1
    for col_idx, label in enumerate(["Fournisseur", "Nb factures", "Total HT", "Total TVA", "Total TTC"], start=1):
        c = ws.cell(row=summary_header_row, column=col_idx, value=label)
        c.font = sum_hdr_font
        c.fill = sum_hdr_fill
        c.alignment = center if col_idx > 1 else left

    ws.row_dimensions[summary_header_row].height = 18

    grand_ht = grand_tva = grand_ttc = 0.0
    grand_count = 0
    has_any_amounts = any(s["has_amounts"] for s in supplier_totals.values())

    for i, (sup_name, st) in enumerate(sorted(supplier_totals.items())):
        r = summary_header_row + 1 + i
        ws.cell(row=r, column=1, value=sup_name).font = sum_bold_font
        ws.cell(row=r, column=2, value=st["count"]).font = sum_row_font
        ws.cell(row=r, column=2).alignment = center
        ws.cell(row=r, column=3, value=_fmt_amount(st["ht"]) if st["has_amounts"] else "").font = sum_row_font
        ws.cell(row=r, column=3).alignment = right
        ws.cell(row=r, column=4, value=_fmt_amount(st["tva"]) if st["has_amounts"] else "").font = sum_row_font
        ws.cell(row=r, column=4).alignment = right
        ws.cell(row=r, column=5, value=_fmt_amount(st["ttc"]) if st["has_amounts"] else "").font = sum_row_font
        ws.cell(row=r, column=5).alignment = right
        ws.row_dimensions[r].height = 16

        grand_count += st["count"]
        grand_ht    += st["ht"]
        grand_tva   += st["tva"]
        grand_ttc   += st["ttc"]

    total_row = summary_header_row + 1 + len(supplier_totals)
    for col_idx, val in enumerate([
        f"TOTAL — {len(invoices)} facture(s)",
        grand_count,
        _fmt_amount(grand_ht)  if has_any_amounts else "",
        _fmt_amount(grand_tva) if has_any_amounts else "",
        _fmt_amount(grand_ttc) if has_any_amounts else "",
    ], start=1):
        c = ws.cell(row=total_row, column=col_idx, value=val)
        c.font = total_font
        c.fill = total_fill
        c.alignment = center if col_idx > 1 else left
    ws.row_dimensions[total_row].height = 18

    # -----------------------------------------------------------------------
    # Column widths, freeze panes, auto-filter
    # -----------------------------------------------------------------------
    for col_idx, (_, min_w, max_w) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_content = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(invoices) + 2)),
            default=min_w,
        )
        ws.column_dimensions[col_letter].width = min(max(max_content + 2, min_w), max_w)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(invoices) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    logger.info(
        "Built Excel summary for %s: %d invoice(s), %d suppliers, %d bytes",
        month_label, len(invoices), len(supplier_totals), buf.getbuffer().nbytes,
    )
    return buf.read()
